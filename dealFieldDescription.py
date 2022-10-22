import pymysql
from openpyxl import load_workbook

conn = pymysql.connect(host='192.168.5.22',
                       port=3307,
                       user='user_yangll',
                       password='vGxw9jWg',
                       database='gps',
                       charset='utf8')
cursor = conn.cursor()

if __name__ == '__main__':
    # 建立空字典
    result = {}
    # workbook = load_workbook('E:\\data\\17-18年.xlsx')
    workbook = load_workbook('D:\\data01\\comment.xlsx')
    sheets = workbook.get_sheet_names()
    ws = workbook.get_sheet_by_name(sheets[0])

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    num = 0
    print('总行数:%s' % (ws.max_row + 1))
    for i in range(1, ws.max_row + 1):
        num = num + 1
        print('====================')
        if i == 1:
            continue

        gps = ws.cell(row=i, column=1).value
        table = ws.cell(row=i, column=2).value

        print('库名: %s' % gps)
        print('表名:%s' % table)

        # sql = '''
        # select dc.classid from d_class dc left join d_group dg on dg.vehicleGroupId = dc.vehicleGroupId where  dc.className = %s and dg.groupName = %s;
        # '''
        # cursor.execute(sql, [className, groupName])
        # classidData = cursor.fetchone()
        # classid = classidData[0]
        # print('classid===', classid)
        #
        # querySql = '''
        # select dv.vehicleId ,dv.workOrderId from d_vehicle dv left join d_track_info dti on dti.vehicleId = dv.vehicleId where dti.sn = %s;
        # '''
        # cursor.execute(querySql, sn)
        # vehicleAndWorkOrderId = cursor.fetchone()
        # vehicleId = vehicleAndWorkOrderId[0]
        # workOrderId = vehicleAndWorkOrderId[1]
        # print('vehicleId===', vehicleId, 'workOrderId===', workOrderId)
        # updateVehicleSql = '''
        # update `d_vehicle` set `classId` = {} where `vehicleId` = {};
        # '''.format(classid, vehicleId)
        #
        # updateVehicleWorkorderSql = '''
        # update `d_vehicle_workorder` set `classId` = {} where `workOrderId` = {};
        # '''.format(classid, workOrderId)
        # print(updateVehicleSql, updateVehicleWorkorderSql)
        # with open("H:\\PythonWorkSpaces\\importexceldata\\sql03.sql", "a+") as fw:
        #     fw.write(updateVehicleSql.strip())
        #     fw.write("\n")
        #     fw.write(updateVehicleWorkorderSql.strip())
        #     fw.write("\n")

    print("num : %s" % num)
