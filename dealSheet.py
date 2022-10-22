from openpyxl import load_workbook
import connectiontool.connection as Connection

gpsConn = Connection.Mysql(
    host='192.168.5.22',
    port=3307,
    dbname='gps',
    user='user_yangll',
    password='vGxw9jWg')

neshieldConn = Connection.Mysql(
    host='192.168.5.23',
    port=3306,
    dbname='neshield',
    user='user_yangll',
    password='vGxw9jWg')

workbook = load_workbook('D:\\data01\\comment.xlsx')
sheets = workbook.get_sheet_names()


def dealOneSheet(pageNum):
    ws = workbook.get_sheet_by_name(sheets[pageNum])

    num = 0
    print('总行数:%s' % (ws.max_row + 1))
    for i in range(1, ws.max_row + 1):
        num = num + 1
        print('====================')
        if i == 1:
            continue
        gps = ws.cell(row=i, column=1).value
        tableName = ws.cell(row=i, column=2).value
        tableComment = ws.cell(row=i, column=3).value

        print('库名: %s' % gps)
        print('表名:%s' % tableName)
        print('表注释名:%s' % tableComment)
        sql = f'''
        alter table {gps}.{tableName} comment '{tableComment}';
        '''
        writeFile(sql, "gpsSql")
    print("num : %s" % num)


def dealTwoSheet(pageNum):
    ws = workbook.get_sheet_by_name(sheets[pageNum])

    num = 0
    print('总行数:%s' % (ws.max_row + 1))
    for i in range(1, ws.max_row + 1):
        num = num + 1
        print('====================')
        if i == 1:
            continue
        gps = ws.cell(row=i, column=1).value
        tableName = ws.cell(row=i, column=2).value
        columnName = ws.cell(row=i, column=3).value
        columnDefault = ws.cell(row=i, column=4).value
        isNullable = ws.cell(row=i, column=5).value
        columnType = ws.cell(row=i, column=6).value
        extra = ws.cell(row=i, column=7).value
        columnComment = ws.cell(row=i, column=8).value

        print('库名: %s' % gps)
        print('表名:%s' % tableName)
        print('列名:%s' % columnName)
        print('列默认值:%s' % columnDefault)
        print('是否为NULL:%s' % isNullable)
        print('列类型:%s' % columnType)
        print('是否自增:%s' % extra)
        print('列注释名:%s' % columnComment)
        if extra != 'auto_increment':
            if isNullable == "NO":
                isNullable = 'NOT NULL'
            else:
                isNullable = 'NULL'
        else:
            isNullable = ''

        if columnDefault is not None:
            columnDefault = f'DEFAULT {columnDefault}'
        else:
            columnDefault = ''

        if extra is None:
            extra = ''

        sql = f'''
        ALTER TABLE {gps}.{tableName} modify {columnName}  {columnType} {columnDefault} {isNullable} {extra} COMMENT '{columnComment}';
        '''
        writeFile(sql, "gpsSql")
    print("num : %s" % num)


def dealThreeSheet(pageNum):
    ws = workbook.get_sheet_by_name(sheets[pageNum])

    num = 0
    print('总行数:%s' % (ws.max_row + 1))
    for i in range(1, ws.max_row + 1):
        num = num + 1
        print('====================')
        if i == 1:
            continue
        neshield = ws.cell(row=i, column=1).value
        tableName = ws.cell(row=i, column=2).value
        tableComment = ws.cell(row=i, column=3).value

        print('库名: %s' % neshield)
        print('表名:%s' % tableName)
        print('表注释名:%s' % tableComment)
        sql = f'''
        alter table {neshield}.{tableName} comment '{tableComment}';
        '''
        writeFile(sql, "neshieldSql")
    print("num : %s" % num)


def dealFourSheet(pageNum):
    ws = workbook.get_sheet_by_name(sheets[pageNum])

    num = 0
    print('总行数:%s' % (ws.max_row + 1))
    for i in range(1, ws.max_row + 1):
        num = num + 1
        print('====================')
        if i == 1:
            continue
        neshield = ws.cell(row=i, column=1).value
        tableName = ws.cell(row=i, column=2).value
        columnName = ws.cell(row=i, column=3).value
        columnDefault = ws.cell(row=i, column=4).value
        isNullable = ws.cell(row=i, column=5).value
        columnType = ws.cell(row=i, column=6).value
        extra = ws.cell(row=i, column=7).value
        columnComment = ws.cell(row=i, column=8).value

        print('库名: %s' % neshield)
        print('表名:%s' % tableName)
        print('列名:%s' % columnName)
        print('列默认值:%s' % columnDefault)
        print('是否为NULL:%s' % isNullable)
        print('列类型:%s' % columnType)
        print('是否自增:%s' % extra)
        print('列注释名:%s' % columnComment)
        if extra != 'auto_increment':
            if isNullable == "NO":
                isNullable = 'NOT NULL'
            else:
                isNullable = 'NULL'
        else:
            isNullable = ''

        if columnDefault is not None:
            columnDefault = f'DEFAULT {columnDefault}'
        else:
            columnDefault = ''

        if extra is None:
            extra = ''

        sql = f'''
        ALTER TABLE {neshield}.{tableName} modify {columnName}  {columnType} {columnDefault} {isNullable} {extra} COMMENT '{columnComment}';
        '''
        writeFile(sql, "neshieldSql")
    print("num : %s" % num)


def writeFile(sql, fileName):
    with open("H:\\PythonWorkSpaces\\verifyMysqlData\\" + fileName + ".sql", "a+", encoding='utf-8') as fw:
        fw.write(sql.strip())
        fw.write("\n")


if __name__ == '__main__':
    dealOneSheet(0)
    dealTwoSheet(1)
    dealThreeSheet(2)
    dealFourSheet(3)
